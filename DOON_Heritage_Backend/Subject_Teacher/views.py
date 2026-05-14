from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from .models import SubjectTeacherProfile, SubjectTeacherClassAssignment
from .serializers import SubjectTeacherProfileSerializer, SubjectTeacherClassAssignmentSerializer
from Student.models import ClassRoom, Exam, Student, StudentMark, Subject
from Student.serializers import StudentSerializer, StudentMarkSerializer, SubjectSerializer

class IsSubjectTeacher(permissions.BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and hasattr(request.user, 'subject_teacher_profile'))

class MySubjectsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = SubjectTeacherClassAssignmentSerializer
    permission_classes = [IsSubjectTeacher]

    def get_queryset(self):
        return self.request.user.subject_teacher_profile.assignments.select_related('classroom', 'subject').all()

class MySubjectStudentsViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsSubjectTeacher]

    def get_queryset(self):
        class_ids = self.request.user.subject_teacher_profile.assignments.values_list('classroom_id', flat=True).distinct()
        return Student.objects.select_related('classroom').filter(classroom_id__in=class_ids)

class SubjectTeacherMarksViewSet(viewsets.ModelViewSet):
    serializer_class = StudentMarkSerializer
    permission_classes = [IsSubjectTeacher]

    def get_queryset(self):
        subject_ids = self.request.user.subject_teacher_profile.assignments.values_list('subject_id', flat=True).distinct()
        return StudentMark.objects.select_related('student', 'subject', 'exam').filter(subject_id__in=subject_ids)


def _get_authorized_assignment(user, class_name, subject_name):
    return user.subject_teacher_profile.assignments.select_related('classroom', 'subject').filter(
        classroom__name=class_name,
        subject__name=subject_name
    ).first()


def _get_exam_for_class(exam_id, classroom):
    return Exam.objects.filter(id=exam_id, classrooms=classroom).first()


def _safe_decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _excel_value_to_string(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@api_view(['GET'])
@permission_classes([IsSubjectTeacher])
def download_marks_template(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError:
        return Response({'error': 'openpyxl is not installed.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    class_name = request.query_params.get('class_name', '').strip()
    subject_name = request.query_params.get('subject_name', '').strip()
    exam_id = request.query_params.get('exam_id', '').strip()

    if not class_name or not subject_name or not exam_id:
        return Response({'error': 'class_name, subject_name, and exam_id are required.'}, status=status.HTTP_400_BAD_REQUEST)

    assignment = _get_authorized_assignment(request.user, class_name, subject_name)
    if not assignment:
        return Response({'error': 'You are not assigned to this class and subject.'}, status=status.HTTP_403_FORBIDDEN)

    exam = _get_exam_for_class(exam_id, assignment.classroom)
    if not exam:
        return Response({'error': 'This exam is not assigned to the selected class.'}, status=status.HTTP_400_BAD_REQUEST)

    students = Student.objects.select_related('classroom').filter(classroom=assignment.classroom).order_by('roll_number', 'name')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Marks Entry'

    headers = ['Roll', 'Student Name', 'Class', 'Subject', 'Full Marks', 'Marks Obtained']
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.protection = Protection(locked=True)

    for student in students:
        sheet.append([
            student.roll_number or '',
            student.name,
            assignment.classroom.name,
            assignment.subject.name,
            exam.full_marks,
            ''
        ])

    sheet.freeze_panes = 'A2'

    max_row = max(sheet.max_row, 2)
    validation = DataValidation(
        type='decimal',
        operator='between',
        formula1='0',
        formula2=f'$E2',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='Invalid marks',
        error='Marks obtained must be numeric and less than or equal to Full Marks.'
    )
    sheet.add_data_validation(validation)
    validation.add(f'F2:F{max_row}')

    for row in sheet.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.protection = Protection(locked=cell.column != 6)

    sheet.protection.sheet = True
    sheet.protection.password = 'marks'

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(max_length + 2, 12)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{class_name}_{subject_name}_marks.xlsx".replace(' ', '_')
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['POST'])
@permission_classes([IsSubjectTeacher])
def upload_marks_excel(request):
    uploaded_file = request.FILES.get('file')
    exam_id = request.data.get('exam_id')

    if not uploaded_file:
        return Response({'error': 'Excel file is required.'}, status=status.HTTP_400_BAD_REQUEST)
    if not uploaded_file.name.lower().endswith('.xlsx'):
        return Response({'error': 'Only .xlsx files are allowed.'}, status=status.HTTP_400_BAD_REQUEST)
    if not exam_id:
        return Response({'error': 'exam_id is required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import pandas as pd
        dataframe = pd.read_excel(uploaded_file)
    except Exception:
        return Response({'error': 'Invalid or corrupt Excel file.'}, status=status.HTTP_400_BAD_REQUEST)

    required_columns = {'Roll', 'Class', 'Subject', 'Full Marks', 'Marks Obtained'}
    missing_columns = required_columns.difference(dataframe.columns)
    if missing_columns:
        return Response({'error': f"Missing columns: {', '.join(sorted(missing_columns))}"}, status=status.HTTP_400_BAD_REQUEST)

    processed = 0
    saved = 0
    failed_rows = []

    for index, row in dataframe.iterrows():
        processed += 1
        row_number = index + 2

        roll = '' if pd.isna(row.get('Roll')) else _excel_value_to_string(row.get('Roll'))
        class_name = '' if pd.isna(row.get('Class')) else _excel_value_to_string(row.get('Class'))
        subject_name = '' if pd.isna(row.get('Subject')) else _excel_value_to_string(row.get('Subject'))
        full_marks = _safe_decimal(None if pd.isna(row.get('Full Marks')) else row.get('Full Marks'))
        marks = _safe_decimal(None if pd.isna(row.get('Marks Obtained')) else row.get('Marks Obtained'))

        if not roll:
            failed_rows.append({'row': row_number, 'reason': 'Roll is required.'})
            continue
        if not class_name or not subject_name:
            failed_rows.append({'row': row_number, 'reason': 'Class and Subject are required.'})
            continue

        assignment = _get_authorized_assignment(request.user, class_name, subject_name)
        if not assignment:
            failed_rows.append({'row': row_number, 'reason': 'Teacher is not assigned to this class and subject.'})
            continue

        exam = _get_exam_for_class(exam_id, assignment.classroom)
        if not exam:
            failed_rows.append({'row': row_number, 'reason': 'Exam is not assigned to this class.'})
            continue

        if marks is None:
            failed_rows.append({'row': row_number, 'reason': 'Marks Obtained must be numeric.'})
            continue
        if full_marks is None:
            failed_rows.append({'row': row_number, 'reason': 'Full Marks must be numeric.'})
            continue
        if marks < 0:
            failed_rows.append({'row': row_number, 'reason': 'Marks Obtained cannot be negative.'})
            continue
        if marks > full_marks:
            failed_rows.append({'row': row_number, 'reason': 'Marks Obtained cannot exceed Full Marks.'})
            continue

        student = Student.objects.select_related('classroom').filter(
            classroom=assignment.classroom,
            roll_number=roll
        ).first()
        if not student:
            failed_rows.append({'row': row_number, 'reason': f'Student with roll {roll} was not found in {class_name}.'})
            continue

        StudentMark.objects.update_or_create(
            student=student,
            exam=exam,
            subject=assignment.subject,
            defaults={'marks': marks}
        )
        saved += 1

    return Response({
        'processed': processed,
        'saved': saved,
        'failed': len(failed_rows),
        'failed_rows': failed_rows
    })
